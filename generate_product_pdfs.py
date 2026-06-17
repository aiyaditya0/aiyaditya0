import os
import re
import openpyxl
from fpdf import FPDF
from fpdf.enums import XPos, YPos

class BeautifulProductPDF(FPDF):
    def __init__(self):
        super().__init__()
        self.set_margins(20, 20, 20)
        self.set_auto_page_break(True, margin=20)
        
    def header(self):
        # Draw dark background on every page
        self.set_fill_color(5, 5, 5) # #050505
        self.rect(0, 0, self.w, self.h, 'F')

def sanitize_text(text):
    if not text:
        return ""
    text = str(text)
    text = text.replace('↓', ' -> ')
    text = text.replace('→', ' -> ')
    text = text.replace('•', '-')
    text = text.replace('…', '...')
    text = text.replace('“', '"').replace('”', '"')
    text = text.replace('‘', "'").replace('’', "'")
    text = text.replace('₹', 'Rs ')
    return text.encode('latin-1', 'ignore').decode('latin-1')

def make_product_id(name):
    # Strip numbers at start like "1. ", "10. ", etc.
    clean = re.sub(r'^\d+\.\s*', '', name)
    # Lowercase, replace non-alphanumeric with hyphen
    clean = clean.lower()
    clean = re.sub(r'[^a-z0-9\s-]', '', clean)
    clean = re.sub(r'[\s-]+', '-', clean).strip('-')
    return clean

def generate_pdfs():
    excel_path = "PRODUCT.xlsx"
    output_dir = "IMPORTANT ASSET"
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    products_catalog = {}
    generated_count = 0
    
    print("Starting PDF generation from Excel...")
    print("=" * 60)
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        category = row[0]
        name = row[1]
        price = row[2]
        unit = row[3]
        link = row[4]
        
        # Skip empty rows
        if not name or not category:
            continue
            
        category = str(category).strip()
        name = str(name).strip()
        link = str(link).strip() if link else ""
        
        # Handle price string
        price_str = str(price).strip().upper() if price is not None else ""
        is_free = "FREE" in price_str or price == 0 or not price
        
        # Skip free items for PDF generation (links provided directly on front-end)
        if is_free:
            print(f"Skipping Free item: {name}")
            continue
            
        prod_id = make_product_id(name)
        pdf_filename = f"{prod_id}.pdf"
        pdf_path = os.path.join(output_dir, pdf_filename)
        
        print(f"Generating PDF for: {name} (ID: {prod_id})")
        
        pdf = BeautifulProductPDF()
        pdf.add_page()
        
        # 1. Logo
        logo_path = 'primary-logo.webp'
        if os.path.exists(logo_path):
            pdf.image(logo_path, x=45, y=25, w=120)
            pdf.ln(45)
        else:
            pdf.ln(10)
            pdf.set_font('helvetica', 'B', 28)
            pdf.set_text_color(255, 138, 0)
            pdf.cell(0, 15, 'FutureWithAi', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(15)
            
        # 2. Headline
        pdf.set_font('helvetica', 'B', 18)
        pdf.set_text_color(255, 255, 255)
        pdf.multi_cell(0, 10, sanitize_text(f"Thank you for purchasing!\n{name}"), align='C')
        pdf.ln(8)
        
        # 3. Description
        pdf.set_font('helvetica', '', 11)
        pdf.set_text_color(221, 193, 174) # var(--text-secondary)
        instructions = (
            "Your digital product files are ready for download. "
            "Please click the button below to get lifetime access to the secure folder (Google Drive or MEGA).\n\n"
            "We recommend bookmarking the link or downloading the files to your local drive for fast offline access."
        )
        pdf.multi_cell(0, 6, sanitize_text(instructions), align='C')
        pdf.ln(12)
        
        # 4. CTA Download Button
        btn_w = 150
        btn_h = 14
        btn_x = (pdf.w - btn_w) / 2
        pdf.set_x(btn_x)
        
        pdf.set_fill_color(255, 138, 0) # Orange background
        pdf.set_text_color(5, 5, 5) # Dark text
        pdf.set_font('helvetica', 'B', 13)
        
        # Add clickable cell
        pdf.cell(btn_w, btn_h, "CLICK HERE TO DOWNLOAD FILES", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True, link=link)
        pdf.ln(15)
        
        # 5. Footer divider & support details
        pdf.set_draw_color(35, 35, 35)
        pdf.set_line_width(0.5)
        pdf.line(30, pdf.get_y(), pdf.w - 30, pdf.get_y())
        pdf.ln(10)
        
        pdf.set_font('helvetica', '', 10)
        pdf.set_text_color(165, 140, 123) # var(--text-muted)
        support_info = (
            "If you face any issues with the link or downloads, please contact us:\n"
            "WhatsApp Support: +91 70658 15743  |  Email: anshumanenterprises1119@gmail.com\n"
            "Always happy to help you build FutureWithAI!"
        )
        pdf.multi_cell(0, 5.5, sanitize_text(support_info), align='C')
        
        # Save PDF
        pdf.output(pdf_path)
        generated_count += 1
        
        # Save to catalog mapping
        products_catalog[prod_id] = {
            "filename": pdf_filename,
            "displayName": name
        }
        
    print("=" * 60)
    print(f"Successfully generated {generated_count} delivery PDFs.")
    
    # Write Apps Script mapping to a txt file
    catalog_mapping_path = "apps_script_catalog_entries.txt"
    with open(catalog_mapping_path, "w", encoding="utf-8") as f:
        f.write("// Copy this code block and paste it inside PRODUCT_CATALOG in google-apps-script.gs\n")
        f.write("const NEW_PRODUCT_ENTRIES = {\n")
        for pid, info in products_catalog.items():
            f.write(f"  '{pid}': {{\n")
            f.write(f"    filename: '{info['filename']}',\n")
            f.write(f"    displayName: '{info['displayName']}'\n")
            f.write("  },\n")
        f.write("};\n")
        
    print(f"Apps Script mapping block written to '{catalog_mapping_path}'.")

if __name__ == "__main__":
    generate_pdfs()
