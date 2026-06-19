import openpyxl

def main():
    excel_path = "PRODUCT.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    print(f"Reading spreadsheet: {excel_path}")
    print(f"Total rows: {sheet.max_row}")
    
    for row_idx in range(2, sheet.max_row + 1):
        category = sheet.cell(row=row_idx, column=1).value
        name = sheet.cell(row=row_idx, column=2).value
        price = sheet.cell(row=row_idx, column=3).value
        unit = sheet.cell(row=row_idx, column=4).value
        link = sheet.cell(row=row_idx, column=5).value
        
        print(f"Row {row_idx}:")
        print(f"  Category: {category}")
        print(f"  Name: {name}")
        print(f"  Price: {price}")
        print(f"  Unit: {unit}")
        print(f"  Link: {link}")
        print("-" * 30)

if __name__ == "__main__":
    main()
